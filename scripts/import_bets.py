#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_GLOB = "*.xlsx"

COMPACT_PHASE_OFFSETS = {
    "16 avos": 72,
    "round of 32": 72,
    "oitavas": 88,
    "round of 16": 88,
    "quartas": 96,
    "semifinal": 100,
    "semi": 100,
    "3 lugar": 102,
    "terceiro": 102,
    "final": 103,
}

ALIASES = {
    "AFRICA DO SUL": "South Africa",
    "ÁFRICA DO SUL": "South Africa",
    "ALEMANHA": "Germany",
    "ARGELIA": "Argélia",
    "ARGÉLIA": "Argélia",
    "AUSTRIA": "Áustria",
    "BÉLGICA": "Belgium",
    "BELGICA": "Belgium",
    "ARGENTINA": "Argentina",
    "AUSTRÁLIA": "Australia",
    "BRASIL": "Brasil",
    "Bosnia": "Bósnia",
    "Bósnia": "Bósnia",
    "BÓSNIA": "Bósnia",
    "CABO VERDE": "Cape Verde",
    "CANADÁ": "Canada",
    "COLOMBIA": "Colombia",
    "COLÔMBIA": "Colombia",
    "Coréia": "Coreia do Sul",
    "COSTA DO MARFIM": "Côte d'Ivoire",
    "Costa do Marfin": "Costa do Marfim",
    "CROACIA": "Croácia",
    "Egiito": "Egito",
    "Egito": "Egito",
    "EGITO": "Egito",
    "EQUADOR": "Equador",
    "ESPANHA": "Spain",
    "ESTADOS UNIDOS": "USA",
    "FRANCE": "França",
    "France": "França",
    "FRANÇA": "França",
    "GANA": "Ghana",
    "Austria": "Áustria",
    "Holanda": "Países Baixos",
    "HOLANDA": "Netherlands",
    "INGLATERRA": "England",
    "JAPAO": "Japan",
    "JAPÃO": "Japan",
    "MARROCOS": "Morocco",
    "MÉXICO": "Mexico",
    "MEXICO": "Mexico",
    "NORUEGA": "Norway",
    "PARAGUAI": "Paraguay",
    "PORTUGAL": "Portugal",
    "Congo": "RD Congo",
    "CONGO": "RD Congo",
    "Rep Tcheca": "Rep. Tcheca",
    "SENEGAL": "Senegal",
    "SUÉCIA": "Sweden",
    "SUÍCA": "Switzerland",
    "SUÍÇA": "Switzerland",
    "SUIÇA": "Switzerland",
}

PARTICIPANT_ALIASES = {
    "ADAILTON": "Adailton",
    "ARTHUR": "Arthur",
    "CARRION": "Carrion",
    "RICARDO": "Carrion",
    "RICARDO C": "Carrion",
    "RICARDO C.": "Carrion",
    "RICARDO CARRION": "Carrion",
    "EDU": "Eduardo Azevedo",
    "EDUARDO": "Eduardo Azevedo",
    "EDUARDO AZEVEDO": "Eduardo Azevedo",
    "HENRIQUE": "Henrique Glasmeyer",
    "HENRIQUE GLASMEYER": "Henrique Glasmeyer",
    "MARCELO": "Marcelo Bernardino",
    "MARCELO BERNARDINO": "Marcelo Bernardino",
    "RODRIGO COBRA": "Rodrigo Cobra",
    "COBRA": "Rodrigo Cobra",
    "VITOR": "Vitor",
}


def clean_name(value):
    if value is None:
        return ""
    text = str(value).strip()
    return ALIASES.get(text, text)


def normalized_text(value) -> str:
    return unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii").lower()


def canonical_participant_name(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    normalized = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii").upper()
    return PARTICIPANT_ALIASES.get(normalized, text.title() if text.isupper() else text)


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", ascii_only).strip("-").lower()
    return slug or "participante"


def as_score(value) -> int | None:
    if value is None or value == "":
        return None
    return int(value)


def as_int(value) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def participant_name(path: Path, worksheet) -> str:
    invalid_names = {"NOME", "SEU NOME", "SELEÇÃO 1", "SELECAO 1", "SELEÇÃO 2", "SELECAO 2", "X", "GRUPO", "#"}
    for cell in ("H3", "H2", "C1", "I3", "I2"):
        value = worksheet[cell].value
        if value and str(value).strip().upper() not in invalid_names:
            name = canonical_participant_name(str(value).strip())
            if name:
                return name
    if worksheet["H3"].value:
        name = canonical_participant_name(str(worksheet["H3"].value).strip())
        if name and name.upper() not in invalid_names:
            return name
    cell_name = worksheet["I2"].value
    if cell_name and str(cell_name).strip().upper() not in invalid_names:
        return canonical_participant_name(str(cell_name).strip())
    playoff_name = worksheet["A3"].value
    if playoff_name and str(playoff_name).strip().upper() not in invalid_names | {"ESCREVA SEU NOME AQUI"}:
        return canonical_participant_name(str(playoff_name).strip())
    stem = re.sub(r"_Apostas fase grupos(?:\s*\(\d+\))?$", "", path.stem)
    stem = re.sub(r"_Apostas fase 16 avos(?:\s*\(\d+\))?$", "", stem, flags=re.IGNORECASE)
    stem = re.sub(r"_Apostas fase oitavas(?:\s*\(\d+\))?$", "", stem, flags=re.IGNORECASE)
    stem = re.sub(r"_Apostas_OITAVAS(?:\s*\(\d+\))?$", "", stem, flags=re.IGNORECASE)
    stem = re.sub(r"_OITAVAS(?:\s*\(\d+\))?$", "", stem, flags=re.IGNORECASE)
    stem = re.sub(r"_Apostas oitavas(?:\s*\(\d+\))?$", "", stem, flags=re.IGNORECASE)
    stem = re.sub(r"_Apostas_Round(?:32|16)(?:\s*\(\d+\))?$", "", stem, flags=re.IGNORECASE)
    stem = re.sub(r"_resultados_playoffs_copa_2026.*$", "", stem, flags=re.IGNORECASE)
    stem = re.sub(r"^SEU_NOME_", "", stem, flags=re.IGNORECASE)
    file_name = stem.replace("_", " ").strip()
    if file_name.isupper():
        file_name = file_name.title()
    return canonical_participant_name(file_name or stem)


def phase_key(value) -> str:
    text = str(value or "").strip()
    normalized = normalized_text(text)
    if "grupo" in normalized:
        return "Fase de Grupos"
    return "Mata-Mata"


def compact_phase_from_context(path: Path, worksheet) -> tuple[str, int]:
    context = " ".join(
        [
            str(path.parent.name),
            str(path.stem),
            str(worksheet.title),
            " ".join(str(worksheet.cell(row, col).value or "") for row in range(1, 5) for col in range(1, 10)),
        ]
    )
    normalized = normalized_text(context)
    if "quartas" in normalized or "quarter" in normalized:
        return "Quartas", 96
    for marker, offset in COMPACT_PHASE_OFFSETS.items():
        if marker in normalized:
            if marker in {"round of 32"}:
                return "16 avos", offset
            if marker in {"round of 16"}:
                return "Oitavas", offset
            if marker in {"semi"}:
                return "Semifinal", offset
            if marker in {"3 lugar", "terceiro"}:
                return "3º lugar", offset
            return marker.title() if marker != "16 avos" else "16 avos", offset
    return "16 avos", 72


def date_iso(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if value is None:
        return ""
    text = str(value).strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}$", text):
        return text
    match = re.match(r"^(\d{1,2})-([A-Za-z]{3})$", text)
    if match:
        months = {
            "jan": "01", "feb": "02", "mar": "03", "apr": "04", "may": "05", "jun": "06",
            "jul": "07", "aug": "08", "sep": "09", "oct": "10", "nov": "11", "dec": "12",
        }
        month = months.get(match.group(2).lower())
        if month:
            return f"2026-{month}-{int(match.group(1)):02d}"
    return ""


def looks_like_playoff_sheet(worksheet) -> bool:
    title = " ".join(str(worksheet.cell(row, col).value or "") for row in range(1, 5) for col in range(1, 10))
    normalized_title = normalized_text(title)
    if (
        "segunda fase" in normalized_title
        or "round of 32" in normalized_title
        or "round of 16" in normalized_title
        or "quartas" in normalized_title
        or "quarter" in normalized_title
        or "semifinal" in normalized_title
        or "semi" in normalized_title
        or "fase eliminatoria" in normalized_title
        or "oitavas" in normalized_title
        or "mata-mata" in normalized_title
    ):
        return True
    headers = [str(worksheet.cell(4, col).value or "").strip().lower() for col in range(1, 13)]
    joined = " ".join(headers)
    return "jogo fifa" in joined and "gols 1" in joined and "gols 2" in joined


def extract_champion(worksheet) -> str:
    def valid_candidate(value: str) -> str:
        candidate = clean_name(value).strip(" _-")
        normalized_candidate = normalized_text(candidate)
        if not candidate:
            return ""
        if "escreva" in normalized_candidate or "pontuacao" in normalized_candidate or "____" in candidate:
            return ""
        return candidate

    for row in range(1, min(worksheet.max_row or 60, 60) + 1):
        for col in range(1, min(worksheet.max_column or 12, 12) + 1):
            value = worksheet.cell(row, col).value
            normalized = normalized_text(value)
            if normalized == "campeao":
                for next_row in range(row + 1, min(row + 4, worksheet.max_row or row + 4) + 1):
                    candidate = valid_candidate(str(worksheet.cell(next_row, col).value or ""))
                    if candidate:
                        return candidate

    for row in worksheet.iter_rows(min_row=35, max_row=60, values_only=True):
        values = [str(value).strip() for value in row if value is not None and str(value).strip()]
        if not values:
            continue
        joined = " ".join(values)
        normalized = normalized_text(joined)
        if "campe" not in normalized:
            continue
        for value in values:
            if ":" in value:
                candidate = valid_candidate(value.split(":", 1)[1])
                if candidate:
                    return candidate
        if len(values) > 1:
            candidate = valid_candidate(values[-1])
            if candidate:
                return candidate
    return ""


def extract_group_workbook(path: Path, worksheet):
    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook.active
    if worksheet.max_row is None or worksheet.max_column is None:
        worksheet.reset_dimensions()
    matches = []
    bets = []
    current_group = ""

    for row in worksheet.iter_rows(min_row=4, max_row=120, values_only=True):
        row = tuple(row or ())
        if len(row) < 8:
            row = row + (None,) * (8 - len(row))
        group_label, match_no, date_value, team_1, goals_1, _, goals_2, team_2 = row[:8]
        if group_label:
            current_group = str(group_label).replace("Grupo", "").strip()
        if not match_no or not team_1 or not team_2:
            continue

        match_id = int(match_no)
        matches.append(
            {
                "id": match_id,
                "phase": "Fase de Grupos",
                "group": current_group,
                "date": date_iso(date_value),
                "team1": clean_name(team_1),
                "team2": clean_name(team_2),
            }
        )
        bets.append(
            {
                "matchId": match_id,
                "g1": as_score(goals_1),
                "g2": as_score(goals_2),
            }
        )

    name = participant_name(path, worksheet)
    return {
        "participant": {
            "id": slugify(name),
            "name": name,
            "file": str(path.name),
            "bets": bets,
        },
        "matches": matches,
    }


def extract_playoff_workbook(path: Path, worksheet):
    matches = []
    bets = []
    compact_format = not any(
        "jogo fifa" in str(worksheet.cell(4, col).value or "").strip().lower()
        for col in range(1, 13)
    )
    compact_phase, compact_offset = compact_phase_from_context(path, worksheet)
    start_row = 4 if compact_format else 5
    end_row = min(worksheet.max_row or 120, 120) if compact_format else min(worksheet.max_row or 120, 120)
    for row in worksheet.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
        row = tuple(row or ())
        if len(row) < 12:
            row = row + (None,) * (12 - len(row))
        if compact_format:
            sequence_no, date_value, team_1, goals_1, _, goals_2, team_2 = row[:7]
            if str(sequence_no or "").strip() == "#":
                continue
            if not sequence_no or not team_1 or not team_2:
                continue
            sequence_no = as_int(sequence_no)
            if sequence_no is None:
                continue
            match_id = sequence_no if sequence_no >= 73 else sequence_no + compact_offset
            phase = compact_phase
        else:
            phase, match_no, date_value, _time_value, _stadium, team_1, goals_1, goals_2, team_2 = row[:9]
            if not match_no or not team_1 or not team_2:
                continue
            match_id = as_int(match_no)
            if match_id is None:
                continue
        matches.append(
            {
                "id": match_id,
                "phase": phase_key(phase),
                "group": "",
                "date": date_iso(date_value),
                "team1": clean_name(team_1),
                "team2": clean_name(team_2),
            }
        )
        bets.append(
            {
                "matchId": match_id,
                "g1": as_score(goals_1),
                "g2": as_score(goals_2),
            }
        )

    name = participant_name(path, worksheet)
    participant = {
        "id": slugify(name),
        "name": name,
        "file": str(path.name),
        "bets": bets,
        "champion": extract_champion(worksheet),
    }
    return {"participant": participant, "matches": matches}


def extract_workbook(path: Path):
    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook.active
    if worksheet.max_row is None or worksheet.max_column is None:
        worksheet.reset_dimensions()
    if looks_like_playoff_sheet(worksheet):
        return extract_playoff_workbook(path, worksheet)
    return extract_group_workbook(path, worksheet)


def dedupe_id(participant, existing):
    base = participant["id"]
    candidate = base
    counter = 2
    while candidate in existing:
        candidate = f"{base}-{counter}"
        counter += 1
    participant["id"] = candidate


def load_base_data(path: Path | None):
    if not path or not path.exists():
        return None
    text = path.read_text(encoding="utf-8")
    match = re.search(r"window\.BOLAO_DATA\s*=\s*(\{.*\});?\s*$", text, flags=re.S)
    if not match:
        return None
    return json.loads(match.group(1))


def build_data(input_dir: Path, base_data=None):
    files = sorted(
        path for path in input_dir.rglob(WORKBOOK_GLOB)
        if not path.name.startswith("~$")
    )
    input_file_names = {path.name for path in files}
    if not files:
        raise SystemExit(f"Nenhum arquivo encontrado em {input_dir} com padrão {WORKBOOK_GLOB!r}.")

    participants_by_id = {}
    matches_by_id = {}
    warnings = []

    if base_data:
        for match in base_data.get("matches", []):
            matches_by_id[int(match["id"])] = {
                **match,
                "team1": clean_name(match.get("team1")),
                "team2": clean_name(match.get("team2")),
            }
        for participant in base_data.get("participants", []):
            preserved_files = []
            for file_name in str(participant.get("file", "")).split(";"):
                file_name = file_name.strip()
                if not file_name or file_name in input_file_names or file_name in preserved_files:
                    continue
                preserved_files.append(file_name)
            participants_by_id[participant["id"]] = {
                **participant,
                "file": "; ".join(preserved_files),
                "bets": list(participant.get("bets", [])),
            }

    for path in files:
        extracted = extract_workbook(path)
        participant = extracted["participant"]
        participant_id = participant["id"]
        if participant_id not in participants_by_id:
            existing_ids = set(participants_by_id)
            dedupe_id(participant, existing_ids)
            participant_id = participant["id"]
            participants_by_id[participant_id] = {
                "id": participant["id"],
                "name": participant["name"],
                "file": participant["file"],
                "bets": [],
            }
        target = participants_by_id[participant_id]
        existing_bets = {bet["matchId"]: bet for bet in target["bets"]}
        for bet in participant["bets"]:
            existing_bets[bet["matchId"]] = bet
        target["bets"] = [existing_bets[key] for key in sorted(existing_bets)]
        if participant.get("champion"):
            target["champion"] = participant["champion"]
        current_files = [file.strip() for file in str(target.get("file", "")).split(";") if file.strip()]
        if participant["file"] not in current_files:
            current_files.append(participant["file"])
        target["file"] = "; ".join(current_files)

        for match in extracted["matches"]:
            existing = matches_by_id.get(match["id"])
            if existing and (existing["team1"], existing["team2"]) != (match["team1"], match["team2"]):
                warnings.append(f"{path.name}: jogo {match['id']} difere da base já importada.")
            matches_by_id[match["id"]] = existing or match

    participants = sorted(participants_by_id.values(), key=lambda participant: participant["name"].lower())
    canonical_matches = [matches_by_id[key] for key in sorted(matches_by_id)]

    return {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sourceFolder": str(input_dir),
        "rules": {
            "exactScorePoints": 10,
            "winnerPoints": 7,
            "drawPoints": 5,
            "wrongPoints": 0,
            "maxPerMatch": 10,
            "stageWeights": {
                "Fase de Grupos": 40,
                "Mata-Mata": 50,
                "Campeão da Copa": 10,
            },
            "stageMaxPoints": {
                "Fase de Grupos": 720,
                "Mata-Mata": 320,
            },
            "championBonusPoints": 10,
        },
        "matches": canonical_matches or [],
        "participants": participants,
        "warnings": warnings,
    }


def main():
    parser = argparse.ArgumentParser(description="Importa apostas do bolão para o dashboard HTML.")
    parser.add_argument(
        "--input",
        default="apostas",
        help="Pasta com arquivos *_Apostas fase grupos.xlsx.",
    )
    parser.add_argument(
        "--output",
        default="data/bolao-data.js",
        help="Arquivo JS gerado para uso pelo dashboard.",
    )
    parser.add_argument(
        "--base",
        default="data/bolao-data.js",
        help="Arquivo JS atual usado como base para preservar fases já importadas.",
    )
    args = parser.parse_args()

    base_path = Path(args.base).expanduser().resolve() if args.base else None
    data = build_data(Path(args.input).expanduser().resolve(), load_base_data(base_path))
    output = Path(args.output).resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    js = "window.BOLAO_DATA = " + json.dumps(data, ensure_ascii=False, indent=2) + ";\n"
    output.write_text(js, encoding="utf-8")
    print(f"Importados {len(data['participants'])} participante(s), {len(data['matches'])} jogo(s).")
    print(f"Dados salvos em {output}")
    if data["warnings"]:
        print("Avisos:")
        for warning in data["warnings"]:
            print(f"- {warning}")


if __name__ == "__main__":
    main()
